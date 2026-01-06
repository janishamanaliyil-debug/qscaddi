import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import re
from difflib import SequenceMatcher


def similar(a, b):
    """Return similarity ratio of two strings with aggressive normalization"""
    # Remove all spaces, dots, punctuation, convert to lowercase
    clean_a = re.sub(r'[^a-z0-9]', '', a.lower())
    clean_b = re.sub(r'[^a-z0-9]', '', b.lower())
    return SequenceMatcher(None, clean_a, clean_b).ratio()


def normalize_text(text):
    """Normalize text for flexible matching"""
    if not text:
        return ""
    text = str(text).strip().upper()
    # Remove extra spaces
    text = re.sub(r'\s+', ' ', text)
    return text


def is_column_type(text, column_type):
    """Check if text matches a column type with flexible matching"""
    text_norm = normalize_text(text)

    patterns = {
        'ITEM': [r'\bITEM\b', r'\bNO\b', r'\bS\.?N\b', r'\bSR\.?NO\b', r'\bSL\.?NO\b'],
        'DESCRIPTION': [r'\bDESCRIPTION\b', r'\bDESC\b', r'\bPARTICULARS\b', r'\bITEM\s*DESC\b', r'\bWORK\s*DESC\b'],
        'QUANTITY': [r'\bQUANTITY\b', r'\bQTY\b', r'\bQNTY\b', r'\bQUAN\b'],
        'UNIT': [r'\bUNIT\b', r'\bUOM\b', r'\bU\.?O\.?M\b'],
        'RATE': [r'\bRATE\b', r'\bUNIT\s*RATE\b', r'\bU\.?RATE\b', r'\bPRICE\b'],
        'AMOUNT': [r'\bAMOUNT\b', r'\bTOTAL\b', r'\bVALUE\b'],
    }

    if column_type not in patterns:
        return False

    # Check if any pattern matches
    for pattern in patterns[column_type]:
        if re.search(pattern, text_norm):
            # For UNIT, make sure it's not UNIT RATE
            if column_type == 'UNIT' and re.search(r'RATE', text_norm):
                continue
            return True
    return False


class TemplateAnalyzer:
    """Analyzes template structure with maximum flexibility"""

    def __init__(self, template_path):
        self.wb = openpyxl.load_workbook(template_path)
        self.structure = {}

    def analyze_template(self, sheet_name):
        """Flexible template analysis that adapts to any structure"""
        ws = self.wb[sheet_name]
        structure = {
            'header_row': None,
            'data_start_row': None,
            'vendor1_name': None,
            'vendor2_name': None,
            'vendor_cells': [],
            'columns': {}
        }

        # Search through first 50 rows for header
        for row_idx in range(1, min(51, ws.max_row + 1)):
            col_map = {}

            for col_idx in range(1, min(30, ws.max_column + 1)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if not cell_value:
                    continue

                # Check for each column type
                if is_column_type(cell_value, 'ITEM'):
                    col_map['ITEM'] = col_idx
                elif is_column_type(cell_value, 'DESCRIPTION'):
                    col_map['DESCRIPTION'] = col_idx
                elif is_column_type(cell_value, 'QUANTITY'):
                    col_map.setdefault('QUANTITY', []).append(col_idx)
                elif is_column_type(cell_value, 'UNIT'):
                    col_map.setdefault('UNIT', []).append(col_idx)
                elif is_column_type(cell_value, 'RATE'):
                    col_map.setdefault('RATE', []).append(col_idx)
                elif is_column_type(cell_value, 'AMOUNT'):
                    col_map.setdefault('AMOUNT', []).append(col_idx)

            # Valid header row must have DESCRIPTION and at least 2 RATEs (for comparison)
            if 'DESCRIPTION' in col_map and len(col_map.get('RATE', [])) >= 2:
                structure['header_row'] = row_idx
                structure['data_start_row'] = row_idx + 1

                rate_cols = col_map['RATE']

                # Search for vendor names in rows above header (up to 10 rows)
                for search_row in range(max(1, row_idx - 10), row_idx):
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=search_row, column=col_idx)
                        if cell.value and str(cell.value).strip():
                            cell_text = str(cell.value).strip()
                            # Skip common non-vendor headers
                            if not any(keyword in cell_text.upper() for keyword in
                                       ['SECTION', 'BILL', 'PAGE', 'SHEET', 'WORKS', 'FITOUT', 'FIT-OUT', 'PROJECT']):
                                structure['vendor_cells'].append({
                                    'row': search_row,
                                    'col': col_idx,
                                    'value': cell_text
                                })

                # Identify vendor columns based on RATE column positions
                vendor1_col_center = rate_cols[0]
                vendor2_col_center = rate_cols[1]

                # Find vendor names closest to rate columns
                for vendor_cell in structure['vendor_cells']:
                    col_idx = vendor_cell['col']
                    # Check if this cell is near vendor 1's rate column
                    if abs(col_idx - vendor1_col_center) <= 3 and not structure['vendor1_name']:
                        structure['vendor1_name'] = vendor_cell['value']
                    # Check if this cell is near vendor 2's rate column
                    elif abs(col_idx - vendor2_col_center) <= 3 and not structure['vendor2_name']:
                        structure['vendor2_name'] = vendor_cell['value']

                structure['columns'] = {
                    'ITEM': col_map.get('ITEM'),
                    'DESCRIPTION': col_map.get('DESCRIPTION'),
                    'VENDOR1': {
                        'QUANTITY': col_map['QUANTITY'][0] if len(col_map.get('QUANTITY', [])) > 0 else None,
                        'UNIT': col_map['UNIT'][0] if len(col_map.get('UNIT', [])) > 0 else None,
                        'RATE': col_map['RATE'][0],
                        'AMOUNT': col_map['AMOUNT'][0] if len(col_map.get('AMOUNT', [])) > 0 else None
                    },
                    'VENDOR2': {
                        'QUANTITY': col_map['QUANTITY'][1] if len(col_map.get('QUANTITY', [])) > 1 else None,
                        'UNIT': col_map['UNIT'][1] if len(col_map.get('UNIT', [])) > 1 else None,
                        'RATE': col_map['RATE'][1],
                        'AMOUNT': col_map['AMOUNT'][1] if len(col_map.get('AMOUNT', [])) > 1 else None
                    }
                }
                break

        return structure

    def get_all_sheet_structures(self):
        all_structures = {}
        for sheet_name in self.wb.sheetnames:
            all_structures[sheet_name] = self.analyze_template(sheet_name)
        return all_structures


class BOQReader:
    """Reads BOQ data with flexible column detection"""

    def __init__(self, boq_path):
        self.wb = openpyxl.load_workbook(boq_path, data_only=True)
        self.boq_name = boq_path.split('\\')[-1].split('/')[-1].replace('.xlsx', '').replace('.xls', '')

    def read_boq_data(self, sheet_name):
        """Flexible BOQ reading that adapts to any column structure"""
        ws = self.wb[sheet_name]
        header_row = None
        cols = {}

        # Search for header row
        for row_idx in range(1, min(51, ws.max_row + 1)):
            temp_cols = {}

            for col_idx in range(1, min(30, ws.max_column + 1)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if not cell_value:
                    continue

                if is_column_type(cell_value, 'ITEM'):
                    temp_cols['ITEM'] = col_idx
                elif is_column_type(cell_value, 'DESCRIPTION'):
                    temp_cols['DESCRIPTION'] = col_idx
                elif is_column_type(cell_value, 'QUANTITY'):
                    temp_cols['QUANTITY'] = col_idx
                elif is_column_type(cell_value, 'UNIT'):
                    temp_cols['UNIT'] = col_idx
                elif is_column_type(cell_value, 'RATE'):
                    temp_cols['RATE'] = col_idx
                elif is_column_type(cell_value, 'AMOUNT'):
                    temp_cols['AMOUNT'] = col_idx

            # Valid if has DESCRIPTION and RATE
            if 'DESCRIPTION' in temp_cols and 'RATE' in temp_cols:
                header_row = row_idx
                cols = temp_cols
                break

        if not header_row:
            return []

        data = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            desc_cell = ws.cell(row=row_idx, column=cols.get('DESCRIPTION'))
            if not desc_cell.value or not str(desc_cell.value).strip():
                continue

            desc_value = str(desc_cell.value).strip()

            # Skip summary/section rows
            skip_keywords = ['CARRIED TO SUMMARY', 'SUMMARY', 'SUB-TOTAL', 'SUB TOTAL',
                             'GRAND TOTAL', 'TOTAL AMOUNT', 'PAGE TOTAL']
            if any(keyword in desc_value.upper() for keyword in skip_keywords):
                continue

            # Skip if it looks like a section header (all caps, short, no numbers)
            if len(desc_value) < 100 and desc_value.isupper() and 'SECTION' in desc_value:
                continue

            row_data = {
                'item': ws.cell(row=row_idx, column=cols.get('ITEM')).value if cols.get('ITEM') else None,
                'description': desc_value,
                'quantity': ws.cell(row=row_idx, column=cols.get('QUANTITY')).value if cols.get('QUANTITY') else None,
                'unit': ws.cell(row=row_idx, column=cols.get('UNIT')).value if cols.get('UNIT') else None,
                'rate': ws.cell(row=row_idx, column=cols.get('RATE')).value if cols.get('RATE') else None,
                'amount': ws.cell(row=row_idx, column=cols.get('AMOUNT')).value if cols.get('AMOUNT') else None,
            }
            data.append(row_data)
        return data

    def get_all_sheets(self):
        """Get all sheet names"""
        return self.wb.sheetnames


class BOQMerger:
    """Merges BOQs with maximum flexibility"""

    def __init__(self, template_path):
        self.template_path = template_path
        self.analyzer = TemplateAnalyzer(template_path)
        self.template_structures = self.analyzer.get_all_sheet_structures()
        self.output_wb = openpyxl.load_workbook(template_path)

    def clean_template_formatting(self, ws, start_row, end_row):
        """Clean all fill colors and text from data rows"""
        from openpyxl.cell.cell import MergedCell
        white_fill = PatternFill(fill_type=None)

        for row in range(start_row, end_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if not isinstance(cell, MergedCell):
                    cell.value = None
                    cell.fill = white_fill

    def update_vendor_headers(self, boq1_name, boq2_name):
        """Replace ALL vendor names in headers with BOQ file names"""
        # Collect all vendor names found
        all_vendor_names = set()

        for sheet_name, structure in self.template_structures.items():
            if structure.get('vendor1_name'):
                all_vendor_names.add(structure['vendor1_name'])
            if structure.get('vendor2_name'):
                all_vendor_names.add(structure['vendor2_name'])
            # Add all vendor cells
            for vcell in structure.get('vendor_cells', []):
                all_vendor_names.add(vcell['value'])

        print(f"\n🔍 Found vendor names in template: {all_vendor_names}")
        print(f"   Will replace with: '{boq1_name}' and '{boq2_name}'\n")

        # Sort by length (longest first) to avoid partial replacements
        vendor_names_list = sorted(list(all_vendor_names), key=len, reverse=True)

        # Track which vendor names belong to which position
        vendor1_set = set()
        vendor2_set = set()

        for sheet_name, structure in self.template_structures.items():
            if structure.get('vendor1_name'):
                vendor1_set.add(structure['vendor1_name'])
            if structure.get('vendor2_name'):
                vendor2_set.add(structure['vendor2_name'])

        # Replace in ALL sheets
        for sheet_name in self.output_wb.sheetnames:
            ws = self.output_wb[sheet_name]

            # Search in first 30 rows
            for row_idx in range(1, min(31, ws.max_row + 1)):
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if not cell.value:
                        continue

                    original_value = str(cell.value).strip()

                    # Skip if merged cell (not top-left)
                    is_top_left = True
                    for merged in list(ws.merged_cells.ranges):
                        if merged.min_row <= row_idx <= merged.max_row and \
                                merged.min_col <= col_idx <= merged.max_col:
                            if row_idx != merged.min_row or col_idx != merged.min_col:
                                is_top_left = False
                            break

                    if not is_top_left:
                        continue

                    new_value = original_value
                    replaced = False

                    # Try exact match first
                    if original_value in vendor1_set:
                        new_value = boq1_name
                        replaced = True
                    elif original_value in vendor2_set:
                        new_value = boq2_name
                        replaced = True
                    else:
                        # Try substring replacement (for compound headers)
                        for vname in vendor_names_list:
                            if vname in original_value:
                                if vname in vendor1_set:
                                    new_value = new_value.replace(vname, boq1_name)
                                    replaced = True
                                elif vname in vendor2_set:
                                    new_value = new_value.replace(vname, boq2_name)
                                    replaced = True

                    if replaced and new_value != original_value:
                        cell.value = new_value
                        print(
                            f"  ✓ {sheet_name}!{get_column_letter(col_idx)}{row_idx}: '{original_value}' → '{new_value}'")

    def find_matching_sheet(self, boq_sheets, target_sheet):
        """Find matching sheet with fuzzy matching"""
        best_match = None
        best_score = 0

        for boq_sheet in boq_sheets:
            score = similar(boq_sheet, target_sheet)
            if score > best_score:
                best_score = score
                best_match = boq_sheet

        if best_score > 0.6:
            return best_match
        return None

    def is_summary_sheet(self, sheet_name):
        """Check if sheet is a summary sheet"""
        summary_keywords = ['SUMMARY', 'TOTAL', 'GRAND']
        sheet_upper = sheet_name.upper()
        return any(keyword in sheet_upper for keyword in summary_keywords)

    def check_content_similarity(self, boq1_data, boq2_data, threshold=0.7):
        """Check if BOQ contents are similar enough to merge"""
        if not boq1_data or not boq2_data:
            return False

        desc1 = [item['description'] for item in boq1_data[:min(10, len(boq1_data))]]
        desc2 = [item['description'] for item in boq2_data[:min(10, len(boq2_data))]]

        match_count = 0
        for d1 in desc1:
            for d2 in desc2:
                if similar(d1, d2) > 0.8:
                    match_count += 1
                    break

        similarity = match_count / max(len(desc1), len(desc2)) if max(len(desc1), len(desc2)) > 0 else 0
        return similarity >= threshold

    def copy_sheet_as_is(self, source_wb, source_sheet_name, target_name):
        """Copy sheet as-is with all formatting"""
        source_ws = source_wb[source_sheet_name]

        if target_name in self.output_wb.sheetnames:
            del self.output_wb[target_name]

        target_ws = self.output_wb.create_sheet(title=target_name)

        for row in source_ws.iter_rows():
            for cell in row:
                target_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    target_cell.font = cell.font.copy()
                    target_cell.border = cell.border.copy()
                    target_cell.fill = cell.fill.copy()
                    target_cell.number_format = cell.number_format
                    target_cell.alignment = cell.alignment.copy()

        for merged in source_ws.merged_cells.ranges:
            target_ws.merge_cells(str(merged))

        for col in source_ws.column_dimensions:
            target_ws.column_dimensions[col].width = source_ws.column_dimensions[col].width

        for row in source_ws.row_dimensions:
            target_ws.row_dimensions[row].height = source_ws.row_dimensions[row].height

    def merge_two_boqs(self, boq1_path, boq2_path):
        boq1_reader = BOQReader(boq1_path)
        boq2_reader = BOQReader(boq2_path)
        boq1_name = boq1_reader.boq_name
        boq2_name = boq2_reader.boq_name

        print(f"📦 Merging: {boq1_name} vs {boq2_name}")

        self.update_vendor_headers(boq1_name, boq2_name)

        boq1_sheets = boq1_reader.get_all_sheets()
        boq2_sheets = boq2_reader.get_all_sheets()

        processed_boq1_sheets = set()
        processed_boq2_sheets = set()

        for template_sheet in self.template_structures.keys():
            # Handle summary sheets separately
            if self.is_summary_sheet(template_sheet):
                print(f"📊 Summary sheet: '{template_sheet}' - copying separately")

                matched_sheet1 = self.find_matching_sheet(boq1_sheets, template_sheet)
                matched_sheet2 = self.find_matching_sheet(boq2_sheets, template_sheet)

                if matched_sheet1:
                    processed_boq1_sheets.add(matched_sheet1)
                    target_name = f"{template_sheet}_{boq1_name}"
                    self.copy_sheet_as_is(boq1_reader.wb, matched_sheet1, target_name)
                    print(f"  ✓ Copied as '{target_name}'")

                if matched_sheet2:
                    processed_boq2_sheets.add(matched_sheet2)
                    target_name = f"{template_sheet}_{boq2_name}"
                    self.copy_sheet_as_is(boq2_reader.wb, matched_sheet2, target_name)
                    print(f"  ✓ Copied as '{target_name}'")

                continue

            matched_sheet1 = self.find_matching_sheet(boq1_sheets, template_sheet)
            matched_sheet2 = self.find_matching_sheet(boq2_sheets, template_sheet)

            if not matched_sheet1 and not matched_sheet2:
                continue

            structure = self.template_structures[template_sheet]

            if not structure['header_row']:
                continue

            ws = self.output_wb[template_sheet]

            boq1_data = boq1_reader.read_boq_data(matched_sheet1) if matched_sheet1 else []
            boq2_data = boq2_reader.read_boq_data(matched_sheet2) if matched_sheet2 else []

            if matched_sheet1:
                processed_boq1_sheets.add(matched_sheet1)
                print(f"✓ '{template_sheet}' ← '{matched_sheet1}' ({boq1_name})")
            if matched_sheet2:
                processed_boq2_sheets.add(matched_sheet2)
                print(f"✓ '{template_sheet}' ← '{matched_sheet2}' ({boq2_name})")

            # Check content similarity
            if boq1_data and boq2_data:
                content_similar = self.check_content_similarity(boq1_data, boq2_data)
                if not content_similar:
                    print(f"⚠️  Content differs - copying separately")
                    if matched_sheet1:
                        self.copy_sheet_as_is(boq1_reader.wb, matched_sheet1, f"{template_sheet}_{boq1_name}")
                    if matched_sheet2:
                        self.copy_sheet_as_is(boq2_reader.wb, matched_sheet2, f"{template_sheet}_{boq2_name}")
                    continue

            if structure['data_start_row']:
                self.clean_template_formatting(ws, structure['data_start_row'], ws.max_row)

            # Merge data
            all_descriptions = []
            desc_to_boq1 = {}
            desc_to_boq2 = {}

            for item in boq1_data:
                desc = item['description']
                all_descriptions.append(desc)
                desc_to_boq1[desc] = item

            for item in boq2_data:
                desc = item['description']
                found = False
                for existing_desc in all_descriptions:
                    if similar(desc, existing_desc) > 0.85:
                        desc_to_boq2[existing_desc] = item
                        found = True
                        break
                if not found:
                    all_descriptions.append(desc)
                    desc_to_boq2[desc] = item

            current_row = structure['data_start_row']

            for idx, description in enumerate(all_descriptions):
                boq1_item = desc_to_boq1.get(description)
                boq2_item = desc_to_boq2.get(description)

                if structure['columns'].get('ITEM'):
                    ws.cell(row=current_row, column=structure['columns']['ITEM']).value = idx + 1
                if structure['columns'].get('DESCRIPTION'):
                    ws.cell(row=current_row, column=structure['columns']['DESCRIPTION']).value = description

                if boq1_item:
                    v1_cols = structure['columns']['VENDOR1']
                    if v1_cols['QUANTITY'] and boq1_item['quantity'] is not None:
                        ws.cell(row=current_row, column=v1_cols['QUANTITY']).value = boq1_item['quantity']
                    if v1_cols['UNIT'] and boq1_item['unit'] is not None:
                        ws.cell(row=current_row, column=v1_cols['UNIT']).value = boq1_item['unit']
                    if v1_cols['RATE'] and boq1_item['rate'] is not None:
                        ws.cell(row=current_row, column=v1_cols['RATE']).value = boq1_item['rate']
                    if v1_cols['AMOUNT'] and boq1_item['amount'] is not None:
                        ws.cell(row=current_row, column=v1_cols['AMOUNT']).value = boq1_item['amount']

                if boq2_item:
                    v2_cols = structure['columns']['VENDOR2']
                    if v2_cols['QUANTITY'] and boq2_item['quantity'] is not None:
                        ws.cell(row=current_row, column=v2_cols['QUANTITY']).value = boq2_item['quantity']
                    if v2_cols['UNIT'] and boq2_item['unit'] is not None:
                        ws.cell(row=current_row, column=v2_cols['UNIT']).value = boq2_item['unit']
                    if v2_cols['RATE'] and boq2_item['rate'] is not None:
                        ws.cell(row=current_row, column=v2_cols['RATE']).value = boq2_item['rate']
                    if v2_cols['AMOUNT'] and boq2_item['amount'] is not None:
                        ws.cell(row=current_row, column=v2_cols['AMOUNT']).value = boq2_item['amount']

                current_row += 1

            print(f"  ✅ Merged {len(all_descriptions)} items")

        # Copy unmatched sheets
        for sheet in boq1_sheets:
            if sheet not in processed_boq1_sheets:
                print(f"📄 Unmatched: '{sheet}' from {boq1_name}")
                self.copy_sheet_as_is(boq1_reader.wb, sheet, f"{sheet}_{boq1_name}")

        for sheet in boq2_sheets:
            if sheet not in processed_boq2_sheets:
                print(f"📄 Unmatched: '{sheet}' from {boq2_name}")
                self.copy_sheet_as_is(boq2_reader.wb, sheet, f"{sheet}_{boq2_name}")

    def add_comparison_column(self):
        """Add color coding for comparison"""
        green_font = Font(color="006100")
        red_font = Font(color="9C0006")

        for sheet_name, structure in self.template_structures.items():
            ws = self.output_wb[sheet_name]
            if not structure['header_row']:
                continue

            v1_rate_col = structure['columns']['VENDOR1']['RATE']
            v2_rate_col = structure['columns']['VENDOR2']['RATE']
            v1_amount_col = structure['columns']['VENDOR1']['AMOUNT']
            v2_amount_col = structure['columns']['VENDOR2']['AMOUNT']

            for row_idx in range(structure['data_start_row'], ws.max_row + 1):
                desc_cell = ws.cell(row=row_idx, column=structure['columns']['DESCRIPTION'])
                if not desc_cell.value or not str(desc_cell.value).strip():
                    continue

                v1_rate_cell = ws.cell(row=row_idx, column=v1_rate_col)
                v2_rate_cell = ws.cell(row=row_idx, column=v2_rate_col)
                if v1_rate_cell.value and v2_rate_cell.value:
                    try:
                        v1_rate = float(v1_rate_cell.value)
                        v2_rate = float(v2_rate_cell.value)
                        if v1_rate < v2_rate:
                            v1_rate_cell.font = green_font
                            v2_rate_cell.font = red_font
                        elif v2_rate < v1_rate:
                            v2_rate_cell.font = green_font
                            v1_rate_cell.font = red_font
                    except (ValueError, TypeError):
                        pass

                if v1_amount_col and v2_amount_col:
                    v1_amount_cell = ws.cell(row=row_idx, column=v1_amount_col)
                    v2_amount_cell = ws.cell(row=row_idx, column=v2_amount_col)
                    if v1_amount_cell.value and v2_amount_cell.value:
                        try:
                            v1_amount = float(v1_amount_cell.value)
                            v2_amount = float(v2_amount_cell.value)
                            if v1_amount < v2_amount:
                                v1_amount_cell.font = green_font
                                v2_amount_cell.font = red_font
                            elif v2_amount < v1_amount:
                                v2_amount_cell.font = green_font
                                v1_amount_cell.font = red_font
                        except (ValueError, TypeError):
                            pass

    def save(self, output_path):
        self.add_comparison_column()
        self.output_wb.save(output_path)
        self.output_wb.close()
        print(f"\n✅ Saved: {output_path}")


# ================= MAIN =================

if __name__ == "__main__":
    template_file = r"C:\Users\Lenovo 32\Desktop\tender_test_boq\tender\Appendix B - Carrefour Sava at Sport Society, Dubai, UAE-.xlsx"
    boq1_file = r"C:\Users\Lenovo 32\Desktop\tender_test_boq\tender\pro.xlsx"
    boq2_file = r"C:\Users\Lenovo 32\Desktop\tender_test_boq\tender\R0.xlsx"
    output_file = r"C:\Users\Lenovo 32\Desktop\tender_test_boq\tender\Output_Comparison.xlsx"

    print("=" * 60)
    print("BOQ MERGER - Enhanced Version")
    print("=" * 60)

    merger = BOQMerger(template_file)
    merger.merge_two_boqs(boq1_file, boq2_file)
    merger.save(output_file)

    print("=" * 60)
    print("✅ MERGE COMPLETE!")
    print("=" * 60)